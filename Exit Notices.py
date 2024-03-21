import json
import os
import glob
from datetime import datetime
from dateutil.relativedelta import relativedelta
import pandas as pd
from openpyxl import load_workbook, utils
from docxtpl import DocxTemplate
import sys
import calendar
import comtypes.client
import subprocess
import time
import tempfile

# Load the configuration settings from the JSON file
with open('config.json') as f:
    config = json.load(f)
    
# Check command-line arguments for date range and CSV file path
if len(sys.argv) != 4:
    print("Usage: python enn.py <start_date> <end_date> <csv_file_path> with dates in YYYY-MM-DD format.")
    sys.exit(1)

csv_file_path = sys.argv[1]
start_date = datetime.strptime(sys.argv[2], "%Y-%m-%d")
end_date = datetime.strptime(sys.argv[3], "%Y-%m-%d")
print("Start Date:", start_date, "End Date:", end_date, "CSV File Path:", csv_file_path)  # Debugging

# Specify the path to the notices folder
notices_folder = config['exit_notices_path']['output_path']

# Use a wildcard (*) to match any Word documents in the folder
word_docs = glob.glob(os.path.join(notices_folder, "*.docx"))

# Loop over the list of matched files and remove each one
for doc in word_docs:
    os.remove(doc)

def convert_csv_to_xlsx(csv_file_path, xlsx_file_path):
    """Converts a CSV file to an Excel XLSX file."""
    df = pd.read_csv(csv_file_path)
    df.to_excel(xlsx_file_path, index=False)

# Create a temporary file
temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)

# Get the path of the temporary file
xlsx_file_path = temp_file.name

# Don't forget to close the temporary file when you're done with it
temp_file.close()

# Now you can use xlsx_file_path as the path to the .xlsx file    

# Convert CSV to Excel
convert_csv_to_xlsx(csv_file_path, xlsx_file_path)  

def convert_to_pdf(input_file_path, output_file_path):
    # Create a Word application object
    word = comtypes.client.CreateObject('Word.Application')

    # Set the application to be invisible
    word.Visible = False

    # Open the Word document
    doc = word.Documents.Open(input_file_path)

    # Save the document as a PDF
    doc.SaveAs(output_file_path, FileFormat=17)  # 17 represents the PDF format in Word

    # Close the document and quit Word
    doc.Close()
    word.Quit()

# Convert CSV to Excel
convert_csv_to_xlsx(csv_file_path, xlsx_file_path)

# Load the Excel workbook
wb = load_workbook(filename=xlsx_file_path)
sheet = wb.active

# Iterate over rows to find entrance notifications
structured_data = []
for row in sheet.iter_rows(min_row=2):
    cell = row[36]  # Column AK
    cell_value = str(cell.value).strip()  # Remove leading and trailing spaces
    if cell_value == 'None':  # Skip if cell value is None
        continue
    try:
        # Try to parse the date in the "mm/dd/yyyy" format
        entrance_date = datetime.strptime(cell_value, "%m/%d/%Y")
    except ValueError:
        try:
            # If that fails, try to parse the date in the "yyyy-mm-dd" format
            entrance_date = datetime.strptime(cell_value, "%Y-%m-%d")
        except ValueError:
            print(f"Could not parse date in {cell_value} with formats %m/%d/%Y or %Y-%m-%d")
            continue  # Skip this row if the date is not in a recognized format
    if start_date <= entrance_date <= end_date:
        print(f"Detected entrance in row {row[0].row}, column {cell.column_letter}, date: {entrance_date.strftime('%B %Y')}")  # Debugging
        first_name = row[4].value  # Column E
        last_name = row[5].value  # Column F
        dob_value = row[7].value  # Column H
        if dob_value == '0000-00-00':
            print(f"Invalid date of birth in row {row[0].row}, skipping this row")
            continue
        dob = datetime.strptime(dob_value, '%m/%d/%Y').strftime('%m/%d/%Y')
        po_first_name = row[16].value  # Column Q
        po_last_name = row[17].value  # Column R
        case_manager_office = row[18].value # Column S
        gender = row[9].value # Column J

        if gender.lower() == 'male':
            gender_pronoun = 'his'
            gender1_pronoun = 'he'
        elif gender.lower() == 'female':
            gender_pronoun = 'her'
            gender1_pronoun = 'she'
        else:
            gender_pronoun = 'their'
            gender1_pronoun = 'they'  

        structured_data.append({
            'Name': f"{first_name} {last_name}",
            'First_Name' : first_name,
            'entrance_date': entrance_date.strftime("%B %d, %Y"),
            'Parole_Officer': f"{po_first_name} {po_last_name}",
            'DOB': dob,
            'Date': datetime.now().strftime("%B %d, %Y"),
            'gender': gender_pronoun,
            'gender1' : gender1_pronoun,
            'case_manager_office' : case_manager_office
        })  
        
# Define the path to your .docx template file
template_path = './Source Documents/EntranceNotifications.Template.docx'

# Generate documents for each entrance notification within the date range
for data in structured_data:
    try:
        doc = DocxTemplate(template_path)
        doc.render(data)
        sanitized_date = data['entrance_date'].replace('/', '.').replace(' ', '.').replace(',', '')
        # Split the name into first and last name
        first_name, last_name = data['Name'].split(maxsplit=1)
        filename = f"{first_name}{last_name}.EntranceNotification.docx"
        
        case_manager_office = data['case_manager_office']

        office_dir = f"C:\\Users\\vaner\\OneDrive\\Free for Life\\F4L Documents\\Thinking for Change\\Entrance Notifications\\{case_manager_office}"
        if not os.path.exists(office_dir):
            os.makedirs(office_dir)


        # Get the parole officer's name
        po_first_name, po_last_name = data['Parole_Officer'].split(maxsplit=1)
        officer_name = f"{po_first_name} {po_last_name}"

        officer_dir = f"{office_dir}\\{officer_name}"
        if not os.path.exists(officer_dir):
            os.makedirs(officer_dir)

        # Generate the output filename
        output_filename = f"{officer_dir}\\{filename}"

        doc.save(output_filename)
        print(f"Document saved: {output_filename}")
        # Convert the Word document to a PDF
        pdf_filename = filename.replace('.docx', '.pdf')
        pdf_output_filename = output_filename.replace('.docx', '.pdf')
        convert_to_pdf(output_filename, pdf_output_filename)
        print(f"PDF saved: {pdf_output_filename}")

        # Wait for the Word application to fully close the document
        time.sleep(10)  # Wait for 5 seconds

        # Delete the Word document
        try:
            os.remove(output_filename)
            print(f"Word document deleted: {output_filename}")
        except Exception as e:
            print(f"Error deleting document for {data['Name']}: {e}")  # More detailed error handling
            pass
    except Exception as e:
        print(f"Error generating document for {data['Name']}: {e}")  # More detailed error handling

# Specify the path to the notifications folder
notifications_folder = config['entrance_notifications_path']['output_path']

# Use a wildcard (*) to match any Word documents in the folder
word_docs = glob.glob(os.path.join(notifications_folder, "*.docx"))

# Loop over the list of matched files and remove each one
for doc in word_docs:
    os.remove(doc)

os.startfile('C:\\Users\\vaner\\OneDrive\\Free for Life\\F4L Documents\\Thinking for Change\\Entrance Notifications')
