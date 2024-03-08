# still opens 2 windows and freezes seemingly unnecessarily before opening the folder containing the generated PDFs for 2nd time

import json
import os
import glob
from datetime import datetime
from dateutil.relativedelta import relativedelta
import pandas as pd
from docxtpl import DocxTemplate
import sys
import calendar
import comtypes.client
import subprocess
import time
import tempfile
import re
from openpyxl.utils import column_index_from_string
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import psutil

csv_file_path = sys.argv[1]
start_date = datetime.strptime(sys.argv[2], "%Y-%m-%d")
end_date = datetime.strptime(sys.argv[3], "%Y-%m-%d")

def convert_csv_to_xlsx(csv_file_path, xlsx_file_path):
    """Converts a CSV file to an Excel XLSX file."""
    df = pd.read_csv(csv_file_path)
    df.to_excel(xlsx_file_path, index=False)

# Load the configuration settings from the JSON file
with open('config.json') as f:
    config = json.load(f)
    
# Check command-line arguments for date range and CSV file path
if len(sys.argv) != 4:
    print("Usage: python Unexcused Absences.py <start_date> <end_date> <csv_file_path> with dates in YYYY-MM-DD format.")
    sys.exit(1)

print("Start Date:", start_date, "End Date:", end_date, "CSV File Path:", csv_file_path)  # Debugging

# Specify the path to the absences folder
absences_folder = config['unexcused_absences_path']['output_path']

# Use a wildcard (*) to match any Word documents in the folder
word_docs = glob.glob(os.path.join(absences_folder, "*.docx"))

def convert_to_pdf(input_file_path, output_file_path):
    # Create a Word application object
    word = comtypes.client.CreateObject('Word.Application')

    # Set the application to be invisible
    word.Visible = False

    # Open the Word document
    doc = word.Documents.Open(input_file_path)

    # Save the document as a PDF
    doc.SaveAs(output_file_path, FileFormat=17)  # 17 represents the PDF format in Word

    # Close the document and quit Word
    doc.Close()

# Create a temporary file
temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)

# Get the path of the temporary file
xlsx_file_path = temp_file.name

# Close the temporary file
temp_file.close()

# Convert CSV to Excel
convert_csv_to_xlsx(csv_file_path, xlsx_file_path)

# Load the Excel workbook
wb = load_workbook(filename=xlsx_file_path)
sheet = wb.active

# Define the path to your .docx template file
template_path = './Source Documents/UnexcusedAbsences.Template.docx'

# Dynamically find month-year header columns
def find_month_year_headers(sheet):
    headers = {}
    header_columns = ['AO', 'CF', 'DW', 'FN']  # The columns to check
    for col in header_columns:
        headers[col] = set()
        for row in range(2, 501):  # Check rows 2 to 500
            cell = sheet[col + str(row)]
            if cell.value:
                headers[col].add(cell.value)
    print("Identified headers:", headers)  # Debugging
    return headers

# Function to find the relevant month and year for an absence
def get_absence_month_year(cell, headers):
    col_index = column_index_from_string(cell.column_letter)
    if col_index >= column_index_from_string('AP') and col_index <= column_index_from_string('CE'):
        header_cell = sheet['AO' + str(cell.row)]
    elif col_index >= column_index_from_string('CG') and col_index <= column_index_from_string('DV'):
        header_cell = sheet['CF' + str(cell.row)]
    elif col_index >= column_index_from_string('DX') and col_index <= column_index_from_string('FM'):
        header_cell = sheet['DW' + str(cell.row)]
    elif col_index >= column_index_from_string('FO') and col_index <= column_index_from_string('HD'):
        header_cell = sheet['FN' + str(cell.row)]
    else:
        return None
    return datetime.strptime(header_cell.value, "%B%Y")  # Parse the month and year from the header cell
headers = find_month_year_headers(sheet)

structured_data = []

# Iterate over rows to find unexcused absences
for row in sheet.iter_rows(min_row=2):
    for cell in row:
        cell_value = str(cell.value).strip()  # Remove leading and trailing spaces
        if cell_value.endswith("❌"):  # Check if the cell value ends with "❌"
            day_of_month = int(cell_value.split()[0])  # Extract the day of the month
            # Find the relevant month and year for the absence
            month_year = get_absence_month_year(cell, headers)
            if month_year:
                # Adjust the absence_date to represent the actual day of the absence 
                absence_date = month_year.replace(day=day_of_month)
                if start_date <= absence_date <= end_date:
                    print(f"Detected absence in row {row[0].row}, column {cell.column_letter}, date: {absence_date.strftime('%B %Y')}")  # Debugging
                    first_name = row[4].value  # Column E
                    last_name = row[5].value  # Column F
                    dob = datetime.strptime(row[7].value, '%Y-%m-%d').strftime('%m/%d/%Y')  # Column H
                    po_first_name = row[16].value  # Column Q
                    po_last_name = row[17].value  # Column R
                    attendance = row[22].value # Column W
                    missed = row[24].value # Column Y
                    gender = row[9].value  # Column J

                    if gender.lower() == 'male':
                        gender_pronoun = 'He'
                    elif gender.lower() == 'female':
                        gender_pronoun = 'She'
                    else:
                        gender_pronoun = 'They'    

                    structured_data.append({
                        'Name': f"{first_name} {last_name}",
                        'entrance_date': absence_date.strftime("%B %d, %Y"),
                        'Parole_Officer': f"{po_first_name} {po_last_name}",
                        'DOB': dob,
                        'Date': datetime.now().strftime("%B %d, %Y"),
                        'gender': gender_pronoun,
                        'gender1' : gender_pronoun,
                        'attendance': attendance,
                        'missed': missed,
                        'unexcused_date' : absence_date.strftime("%B %d, %Y"), # Add this line
            })

# Generate documents for each unexcused absence within the date range
for data in structured_data:
    try:
        doc = DocxTemplate(template_path)
        doc.render(data)
        sanitized_date = data['unexcused_date'].replace('/', '.').replace(' ', '.').replace(',', '')
        # Split the name into first and last name
        first_name, last_name = data['Name'].split(maxsplit=1)
        filename = f"{last_name}.{first_name}.UnexcusedAbsence.{sanitized_date}.docx"
        
        # Get the parole officer's name
        po_first_name, po_last_name = data['Parole_Officer'].split(maxsplit=1)
        officer_name = f"{po_first_name} {po_last_name}"

        # Create a directory named after the officer if it doesn't exist
        officer_dir = f"{absences_folder}\\{officer_name}"
        if not os.path.exists(officer_dir):
            os.makedirs(officer_dir)

        # Generate the output filename
        output_filename = f"{officer_dir}\\{filename}"
        
        doc.save(output_filename)
        print(f"Document saved: {output_filename}")
        
        # Convert the Word document to a PDF
        pdf_filename = filename.replace('.docx', '.pdf')
        pdf_output_filename = output_filename.replace('.docx', '.pdf')
        convert_to_pdf(output_filename, pdf_output_filename)
        print(f"PDF saved: {pdf_output_filename}")

        # Delete the Word document
        try:
            os.remove(output_filename)
            print(f"Word document deleted: {output_filename}")
        except Exception as e:
            print(f"Error deleting document for {data['Name']}: {e}")  # More detailed error handling
            pass
    except Exception as e:
        print(f"Error generating document for {data['Name']}: {e}")  # More detailed error handling

# Specify the path to the absences  folder
absences_folder = config['unexcused_absences_path']['output_path']

time.sleep(5)  # Wait for 5 seconds

# Open the folder containing the generated PDFs
os.startfile(absences_folder)  # Opens the folder in the default file explorer